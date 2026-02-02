import pandas as pd
import re
import glob
import os
import time
from typing import Dict
import xlrd
from openpyxl import Workbook
from datetime import datetime

def str_to_bool(s):
    if s.lower() in ['true', 't', 'yes', 'y', '1']:
        return True
    elif s.lower() in ['false', 'f', 'no', 'n', '0']:
        return False
    else:
        raise ValueError("Invalid input: must be 'True' or 'False'")

def convert_x2x(downloadPath):
    got_file = False
    found_status = True  # file is not downloaded so nothing has been found
    # Grab current file name.
    # interupt=1
    while not got_file:
        try:
            currentFiles = glob.glob(downloadPath + "*.xls")
            got_file = True

        except IndexError:
            found_status = False
            print("File has not finished downloading")
            # interupt += 1
            time.sleep(5)
            return found_status

    for xls_file in currentFiles:
        # Extract the base filename (without extension)
        base_filename = os.path.splitext(os.path.basename(xls_file))[0]

        # Create the corresponding .xlsx filename
        xlsx_file = os.path.join(downloadPath, base_filename + ".xlsx")

        # Check if the destination file already exists
        # counter = 1
        # while os.path.exists(xlsx_file):
        #     # If it exists, append a counter to the file name
        #     newFileName_with_counter = f"{base_filename} ({counter})"
        #     xlsx_file = os.path.join(downloadPath, newFileName_with_counter + ".xlsx")
        #     counter += 1
        if os.path.exists(xlsx_file):
            continue

        # Convert .xls to .xlsx without pandas to support xlrd==1.2.0
        try:
            book = xlrd.open_workbook(xls_file, formatting_info=False)
            wb_out = Workbook()
            first_sheet = True
            for sheet in book.sheets():
                ws_out = wb_out.active if first_sheet else wb_out.create_sheet()
                ws_out.title = (sheet.name or "Sheet1")[:31]
                first_sheet = False

                for row_idx in range(sheet.nrows):
                    for col_idx in range(sheet.ncols):
                        cell = sheet.cell(row_idx, col_idx)
                        value = cell.value
                        if cell.ctype == xlrd.XL_CELL_DATE:
                            try:
                                y, m, d, hh, mm, ss = xlrd.xldate_as_tuple(value, book.datemode)
                                value = datetime(y, m, d, hh, mm, ss)
                            except Exception:
                                pass
                        ws_out.cell(row=row_idx + 1, column=col_idx + 1, value=value)

            wb_out.save(xlsx_file)
            print(f"Converted \"{xls_file}\" to \"{xlsx_file}\"")
        except Exception as e:
            print(f"Failed to convert \"{xls_file}\": {e}")
    return found_status

def extract_product_name(file_name):
    # match = re.search(r'\d+\.\s(.+?)\s\d+\s\d+', file_name)
    # Search for the pattern and capture the first digit and the desired text
    match = re.search(r'(\d+)\.\s(.+?)\s(\d+(\s\d+)?)', file_name)
    if match:
        try:
            number = match.group(1)

        except:
            number=None
            
        try:
            text = match.group(2)
        except:
            text=None
            
        try:
            time = match.group(3)
        except:
            time=None
        return number, text, time
    else:
        return None, None, None
    
def merge_ingredients(downloadPath, fileExtension='.xlsx', export_data=False):
    currentFiles = glob.glob(downloadPath + "*" + fileExtension)
    dfs = []
    merged_chats = []
    for currentFile in currentFiles:
        number, product_name, time = extract_product_name(currentFile)
        print(currentFile)
        # Read only the desired columns
        df = pd.read_excel(currentFile)
        print(f"File: {currentFile}, Columns: {df.shape[1]}")

        try:
            if export_data == False:
                # Set the header to the second row
                df.columns = ['Date', 'Mã_tờ_khai', 'Công_ty_nhập', 'Công_ty_nhập (TA)', 
                'Địa_chỉ', 'Mã_số_thuế', 'Nhà_cung_cấp','Địa_chỉ_(ncc)',
                'Quốc_gia_xuất_xứ', 'Mã_nước_xuất_khẩu', 'HScode', 'Mô_tả_sản_phẩm', 'Số_lượng',
                'Đơn_vị', 'Khối_lượng', 'Thành_tiền', 'Tiền_tệ', 'Đơn_giá',
                '单价单位', '原单价','Tỷ giá', '进口类型', '进口税率', '进口税额', '税额（越南盾)', 'Điều_kiện_giao_hàng', '付款方式',
                '海关名称', '海关代码', '海关代理代码', 'Cảng xuất', '起运港代码', 'Cảng nhập', '目的港代码', 
                '原产国', '原产国家代码','承运人', '追踪号'] # '起运港': cang xuat, '目的港': 'cang nhap
            else:
                df.columns = ['Date', 'Mã_tờ_khai', 'Công_ty_nhập', 
                'Địa_chỉ', "Nước_nhập_khẩu", 'Nhà_cung_cấp', "Mã_số_thuế", "Nước_xuất_khẩu",
                'HScode', 'Mô_tả_sản_phẩm', 'Số_lượng','Đơn_vị', 'Khối_lượng',
                'Thành_tiền', 'Đơn_giá','Tiền_tệ',
                '出口税额',	'出口税率',	'出口税额单位',	'税额（越南盾)','Điều_kiện_giao_hàng','付款方式','Cảng nhập',
                '海关代码',	"海关代理代码",	"Cảng xuất","起运港代码",	"目的港", "目的港代码", "运输方式", "承运人","追踪号"]
        except ValueError as e:
            print(f"Error processing {currentFile}: {e}")
            print(f"Expected 38 columns (export_data=False) or 32 columns (export_data=True), but got {df.shape[1]}")
            continue
        
        # Drop the first row (0 is not the old column names)
        df = df.drop([0])
        # df.to_csv(newFileName+product_name+fileExtension, index=False, encoding='utf-8-sig')
        if product_name:
            df['Sản phẩm'] = product_name
            merged_chats.append(product_name)
            with open("merged_chats.txt", mode='w', encoding= 'utf-8') as file:
                file.write(", \n".join(merged_chats))
        else:
            print(f"{product_name} not found")
                
        dfs.append(df)
                
    if dfs:
        # Reset index for each DataFrame
        dfs = [df.reset_index(drop=True) for df in dfs]
        
        merged_df = pd.concat(dfs, ignore_index=True)
        # Drop duplicated rows
        merged_df = merged_df.drop_duplicates(keep="first")

            
    else:
        print("No data to concatenate.")
        merged_df = pd.DataFrame()
    return merged_df

# Define a mapping of possible variations for each column
column_mapping = {
    "Mã tờ khai": ['Mã_tờ_khai', "Mã tờ khai"],
    'Day': ['day', 'Ngày', 'Day', 'ngày'],
    'Month': ['month', 'Tháng', 'Month'],
    'Year': ['year', 'Năm', 'Year'],
    'Công_ty_nhập': ['Công ty nhập', 'Công_ty_nhập','công ty nhập','Cong_ty_nhap', 'Công ty nhập '],
    'Công ty nhập gộp': ['Công ty nhập gộp', 'Công_ty_nhập_gộp', 'Công_ty_nhập gộp', 'Công ty nhập (gộp)', 'Công_ty_nhập (gộp)', 'Công ty gộp', 'Cong_ty_nhap (gộp)'],
    'Công_ty_nhập (TA)': ['Công_ty_nhập (TA)', "Công_ty_nhập(TA)", "Công_ty_nhập_(TA)",'Công_ty_nhập(TA', 'Công ty nhập(TA)'],
    'Địa_chỉ': ['Địa_chỉ', 'Địa chỉ', 'Địa_chỉ', 'Địa chỉ công ty nhập'],
    'Mã số thuế': ["Mã_số_thuế", 'Mã số thuế'],
    "Nhà_cung_cấp": ['Nhà_cung_cấp', 'Nhà cung cấp', 'NCC', 'Công ty xuất'],
    'Địa_chỉ_(ncc)': ["Địa_chỉ_(ncc)", "Địa_chỉ(ncc)", "Địa chỉ công ty xuất",'Địa chỉ công ty xuất khẩu'],
    "Quốc_gia_xuất_xứ": ["Quốc_gia_xuất_xứ", "Xuất_xứ","Xuất xứ", 'Nước xuất khẩu'],
    'Thương hiệu': ['Brand', 'Brand ','Thương hiệu', 'Thương_hiệu', 'ten_cong_ty_update'],
    "HScode": ["HSCode","HScode", "HS code", "HS Code", "Hscode"],
    'Mô_tả_sản_phẩm': ['Mô_tả_sản_phẩm', 'Mô tả sản phẩm', 'Mo_ta_san_pham'],
    'Hợp_lệ': ['Hợp lệ', 'Hợp_lệ', 'hợp lệ', 'hợp_lệ','Validation','validation'],
    'updated_Số_lượng': ['Updated_số_lượng', 'Update số lượng ','updated_số_lượng', 'Update_số_lượng', 'Updated_Số_lượng',
                         'update_số_lượng','Updated Số lượng', 'Updated số lượng', 'updated_Số_lượng','Updated_So_luong', 'Update số lượng', 'update_số_lượng'],
    "updated_Đơn_vị": ["updated_Đơn_vị","Updated_Đơn_vị", "updated_Đơn_vị ", 'updated đơn vị'],
    "Số lượng": ["Số lượng", "Số_lượng", 'Số lượng theo tờ khai',],
    "Đơn_vị": ["Đơn_vị","Đơn vị", "Đơn vị ", 'ĐVT số lượng'],
    "Khối_lượng": ["Khối_lượng", "Khối lượng"],
    'Thành_tiền': ['Thành_tiền', 'Thành tiền', 'Giá trị', 'Thành tiền ', '金额', 'Gia tri', 'Thành tiền theo tờ khai'],
    'Tiền_tệ': ['Tiền_tệ',"Tiền tệ", 'ĐVT tiền tệ'],
    'Đơn_giá': ['Đơn giá', 'Đơn_giá', 'đơn giá', 'đơn_giá', '单价', ' Đơn giá', 'Đơn giá ', 'Đơn giá USD/kg', 'Đơn_giá ', ' Đơn_giá'],
    '单价单位': ['单价单位', '单价单位  '],
    '原单价': ['原单价', '原单价    '],
    "Tỷ giá": ['Tỷ_giá', "Tỷ giá", "汇率"],
    '进口类型': ['进口类型', '进口类型	'],
    '进口税率': ['进口税率', '进口税率	'],
    '进口税额': ['进口税额', '进口税额	'],
    '税额（越南盾)': ['税额（越南盾)'],
    "Điều_kiện_giao_hàng": ["Điều_kiện_giao_hàng", 'Điều kiện giao hàng'],
    '付款方式': ['付款方式', '付款方式  '],
    '海关名称': ['海关名称', '海关名称  '],
    '海关代码': ['海关代码', '海关代码  '],
    '海关代理代码': ['海关代理代码', '海关代理代码  '],
    '起运港代码': ['起运港代码', '起运港代码    '],
    "Cảng xuất": ['起运港', 'Cảng xuất', '起运港   ', "起运港	"],
    '目的港代码': ['目的港代码', '目的港代码    '],
    "Cảng nhập": ["目的港", "Cảng", "Cảng nhập", "目的港	"],
    '原产国': ['原产国', '原产国    '],
    '原产国家代码': ['原产国家代码', '原产国家代码  '],
    '承运人': ['承运人', '承运人    '],
    '追踪号': ['追踪号', '追踪号    '],
    # "Thành tiền quy đổi theo usd": ["Thành_tiền_usd"],
    'Phân loại công ty nhập': ['Phân_loại_công_ty_nhập', 'Phân loại công ty nhập', 'Phân loại công ty'],
    'Phân loại thị trường': ['left_MarketClassification','Market_Classification','Market Classification', 'MarketClassification', 'Phân vùng thị trường', 'Phân_loại_thị_trường', 'Phân loại thị trường'],
    'is_duplicate': ['is_duplicate', 'Duplicated_MTK', 'is_delete', 'is_duplicated'],
    'Sản phẩm': ['Sản_phẩm', 'Sản phẩm']
}


def take_ingredient(path_filename):
    df = read_and_rename(path_filename)
    # print(unmatched_columns)
    ingredients = df['Sản phẩm'].unique()
    # Convert it to a list and lower all strings
    ingredients_lower = [ing.lower() if isinstance(ing, str) else ing for ing in ingredients]

    # Now 'ingredients_lower' will contain all strings converted to lowercase while keeping non-string elements unchanged
    # print(ingredients_lower)
    return df, ingredients_lower

def get_key_by_value(dictionary, target_values):
    key_list = []
    for value in target_values:
        for key, values in dictionary.items():
            if value in values:
                key_list.append(key)
    return key_list


def read_and_rename(currentFile):
    # Read only the desired columns
    df = pd.read_excel(currentFile, sheet_name='Data', nrows=1)
    matched_columns = []
    unmatched_columns = ()
    for column, variations in column_mapping.items():
        for var in variations:
            if var in df.columns:
                matched_columns.append(var)
                # # Break out of the inner loop and move to the next iteration of the outer loop
                break
        else:
            unmatched_columns = tuple(set(unmatched_columns + (column,)))

    # print("Columns that are matched: ",matched_columns)
    print("Columns that are unmatched in df: ",unmatched_columns)
    if matched_columns:
        # Extract only the desired columns
        df = pd.read_excel(currentFile, sheet_name='Data')
        # Create a dictionary to map old column names to new column names
        column_rename = dict(zip(matched_columns, get_key_by_value(column_mapping, matched_columns)))

        # Rename the columns in df1 using the dictionary
        df.rename(columns=column_rename, inplace=True)
    else:
        return pd.DataFrame()
    return df
                
                
def merge_new_into_old(newFilePath, existedFilePath="C:/Users/Admin/Documents/Old D disk/download/company BfC/merge cac chat/Thi truong dau vao/tat_ca/", outputfolder="merged_files/"):
    currentFiles = glob.glob(existedFilePath + "*" + ".xlsx")
    
    new_df, ingredients = take_ingredient(newFilePath)
    not_found_product = []
    for currentFile in currentFiles:
        number, product_name, time = extract_product_name(currentFile)
        if product_name:
            if product_name.lower() in ingredients:
                print(f"{number}. {product_name.capitalize()}")

                if os.path.exists(f"{outputfolder}{number}. {product_name.title()} {time}.xlsx"):
                    print("This file existed!")
                    continue
                
                # Check if the desired columns exist in the sheet
                existed_df = read_and_rename(currentFile)
                if existed_df.empty:
                    print("no columns/file found")
                
                if existed_df.columns.duplicated().any():
                    # Handle duplicate column names by renaming them
                    duplicated_columns = existed_df.columns[existed_df.columns.duplicated()].tolist()
                    print("Duplicated columns in new_df_ingre:", duplicated_columns)
                    existed_df.columns = [f'{col}_{i}' if existed_df.columns[i] in existed_df.columns[:i] else col for i, col in enumerate(existed_df.columns)]
                    
                new_df_ingre = new_df[new_df['Sản phẩm'] == product_name].copy()
                unmatched_columns_new_df = set(new_df_ingre.columns) - set(existed_df.columns)
                
                try:
                    unmatched_columns_new_df.remove("is_duplicate")
                except KeyError:
                    print("Item 'is_duplicate' not found in the set.")

                new_df_ingre.drop(list(unmatched_columns_new_df), axis=1, inplace=True)
                
                merged_df = pd.concat([existed_df, new_df_ingre], ignore_index=True)
                
                # Check if the output folder exists, if not, create it
                if not os.path.exists(outputfolder):
                    print("Runheare ")
                    os.makedirs(outputfolder)
                
                else:
                    merged_df.to_excel(f"{outputfolder}{number}. {product_name.title()} {time}.xlsx", index=False, sheet_name="Data")
                    print(f"Merge successfully! {outputfolder}{number}. {product_name.title()} {time}.xlsx")
            else:
                not_found_product.append(product_name)
                print(f"{product_name}: product not found")
                continue
        else:
            print(f"{product_name}: product not found")
            continue
        
        if not_found_product:
            print(f"{not_found_product}: not found")
    