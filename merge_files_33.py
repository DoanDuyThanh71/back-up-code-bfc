import pandas as pd
import re
import glob
import os
import time
from typing import Dict
import xlrd
from openpyxl import Workbook
from datetime import datetime

def str_to_bool(s):
    if s.lower() in ['true', 't', 'yes', 'y', '1']:
        return True
    elif s.lower() in ['false', 'f', 'no', 'n', '0']:
        return False
    else:
        raise ValueError("Invalid input: must be 'True' or 'False'")

def convert_x2x(downloadPath):
    got_file = False
    found_status = True
    while not got_file:
        try:
            currentFiles = glob.glob(downloadPath + "*.xls")
            got_file = True

        except IndexError:
            found_status = False
            print("File has not finished downloading")
            time.sleep(5)
            return found_status

    for xls_file in currentFiles:
        base_filename = os.path.splitext(os.path.basename(xls_file))[0]
        xlsx_file = os.path.join(downloadPath, base_filename + ".xlsx")
        if os.path.exists(xlsx_file):
            continue

        try:
            book = xlrd.open_workbook(xls_file, formatting_info=False)
            wb_out = Workbook()
            first_sheet = True
            for sheet in book.sheets():
                ws_out = wb_out.active if first_sheet else wb_out.create_sheet()
                ws_out.title = (sheet.name or "Sheet1")[:31]
                first_sheet = False

                for row_idx in range(sheet.nrows):
                    for col_idx in range(sheet.ncols):
                        cell = sheet.cell(row_idx, col_idx)
                        value = cell.value
                        if cell.ctype == xlrd.XL_CELL_DATE:
                            try:
                                y, m, d, hh, mm, ss = xlrd.xldate_as_tuple(value, book.datemode)
                                value = datetime(y, m, d, hh, mm, ss)
                            except Exception:
                                pass
                        ws_out.cell(row=row_idx + 1, column=col_idx + 1, value=value)

            wb_out.save(xlsx_file)
            print(f"Converted \"{xls_file}\" to \"{xlsx_file}\"")
        except Exception as e:
            print(f"Failed to convert \"{xls_file}\": {e}")
    return found_status

def extract_product_name(file_name):
    # match = re.search(r'\d+\.\s(.+?)\s\d+\s\d+', file_name)
    # Search for the pattern and capture the first digit and the desired text
    match = re.search(r'(\d+)\.\s(.+?)\s(\d+(\s\d+)?)', file_name)
    if match:
        try:
            number = match.group(1)
        except:
            number=None
        try:
            text = match.group(2)
        except:
            text=None
        try:
            time = match.group(3)
        except:
            time=None
        return number, text, time
    else:
        return None, None, None
    
def merge_ingredients(downloadPath, fileExtension='.xlsx', export_data=False):
    currentFiles = glob.glob(downloadPath + "*" + fileExtension)
    dfs = []
    merged_chats = []
    for currentFile in currentFiles:
        number, product_name, time = extract_product_name(currentFile)
        print(currentFile)
        # Read only the desired columns
        df = pd.read_excel(currentFile)

        if len(df.columns) == 38:
            df.columns = ['Date', 'Mã_tờ_khai', 'Công_ty_nhập', 'Công_ty_nhập (TA)', 
            'Địa_chỉ', 'Mã_số_thuế', 'Nhà_cung_cấp','Địa_chỉ_(ncc)',
            'Quốc_gia_xuất_xứ', 'Mã_nước_xuất_khẩu', 'HScode', 'Mô_tả_sản_phẩm', 'Số_lượng',
            'Đơn_vị', 'Khối_lượng', 'Thành_tiền', 'Tiền_tệ', 'Đơn_giá',
            '单价单位', '原单价','Tỷ giá', '进口类型', '进口税率', '进口税额', '税额（越南盾)', 'Điều_kiện_giao_hàng', '付款方式',
            '海关名称', '海关代码', '海关代理代码', 'Cảng xuất', '起运港代码', 'Cảng nhập', '目的港代码', 
            '原产国', '原产国家代码','承运人', '追踪号']
        elif len(df.columns) == 33:
             df.columns = ['Date', 'Mã_tờ_khai', 'Công_ty_nhập', 
            'Địa_chỉ', "Nước_nhập_khẩu", 'Nhà_cung_cấp', "Nhà_cung_cấp (TA)", "Mã_số_thuế", "Quốc_gia_xuất_xứ",
            'HScode', 'Mô_tả_sản_phẩm', 'Số_lượng','Đơn_vị', 'Khối_lượng',
            'Thành_tiền', 'Đơn_giá','Tiền_tệ',
            '出口税额',	'出口税率',	'出口税额单位',	'税额（越南盾)','Điều_kiện_giao_hàng','付款方式','海关名称',
            '海关代码',	"海关代理代码",	"Cảng xuất","起运港代码",	"Cảng nhập", "目的港代码", "运输方式", "承运人","追踪号"]
        elif len(df.columns) == 32:
            df.columns = ['Date', 'Mã_tờ_khai', 'Công_ty_nhập', 
            'Địa_chỉ', "Nước_nhập_khẩu", 'Nhà_cung_cấp', "Mã_số_thuế", "Nước_xuất_khẩu",
            'HScode', 'Mô_tả_sản_phẩm', 'Số_lượng','Đơn_vị', 'Khối_lượng',
            'Thành_tiền', 'Đơn_giá','Tiền_tệ',
            '出口税额',	'出口税率',	'出口税额单位',	'税额（越南盾)','Điều_kiện_giao_hàng','付款方式','Cảng nhập',
            '海关代码',	"海关代理代码",	"Cảng xuất","起运港代码",	"目的港", "目的港代码", "运输方式", "承运人","追踪号"]
        else:
             print(f"Warning: Unexpected column count {len(df.columns)} in {currentFile}")
             continue
        
        # Drop the first row (0 is not the old column names)
        df = df.drop([0])
        if product_name:
            df['Sản phẩm'] = product_name
            merged_chats.append(product_name)
            with open("merged_chats.txt", mode='w', encoding= 'utf-8') as file:
                file.write(", \n".join(merged_chats))
        else:
            print(f"{product_name} not found")
                
        dfs.append(df)
                
    if dfs:
        dfs = [df.reset_index(drop=True) for df in dfs]
        merged_df = pd.concat(dfs, ignore_index=True)
        merged_df = merged_df.drop_duplicates(keep="first")
    else:
        print("No data to concatenate.")
        merged_df = pd.DataFrame()
    return merged_df
