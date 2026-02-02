import pandas as pd
import glob
import os
import xlrd
from openpyxl import Workbook
from datetime import datetime

download_directory = "./Downloads/"

# Convert XLS to XLSX
def convert_xls_to_xlsx(directory):
    files = glob.glob(os.path.join(directory, "*.xls"))
    print(f"Found {len(files)} xls files.")
    for xls_file in files:
        base_filename = os.path.splitext(os.path.basename(xls_file))[0]
        xlsx_file = os.path.join(directory, base_filename + ".xlsx")
        
        if os.path.exists(xlsx_file):
            print(f"Skipping {xls_file} (xlsx already exists)")
            continue
            
        try:
            print(f"Converting {xls_file}...")
            book = xlrd.open_workbook(xls_file, formatting_info=False)
            wb_out = Workbook()
            first_sheet = True
            for sheet in book.sheets():
                ws_out = wb_out.active if first_sheet else wb_out.create_sheet()
                ws_out.title = (sheet.name or "Sheet1")[:31]
                first_sheet = False

                for row_idx in range(sheet.nrows):
                    for col_idx in range(sheet.ncols):
                        cell = sheet.cell(row_idx, col_idx)
                        value = cell.value
                        if cell.ctype == xlrd.XL_CELL_DATE:
                            try:
                                y, m, d, hh, mm, ss = xlrd.xldate_as_tuple(value, book.datemode)
                                value = datetime(y, m, d, hh, mm, ss)
                            except Exception:
                                pass
                        ws_out.cell(row=row_idx + 1, column=col_idx + 1, value=value)

            wb_out.save(xlsx_file)
            print(f"Converted to {xlsx_file}")
        except Exception as e:
            print(f"Failed to convert {xls_file}: {e}")

convert_xls_to_xlsx(download_directory)

# Inspect headers of the first xlsx file
xlsx_files = glob.glob(os.path.join(download_directory, "*.xlsx"))
if xlsx_files:
    target_file = xlsx_files[0]
    print(f"\nInspecting {target_file}")
    
    # Try reading header at row 0, 1, 2 to see which one makes sense
    for header_row in [0, 1]:
        print(f"\n--- Header Row: {header_row} ---")
        try:
            df = pd.read_excel(target_file, header=header_row, nrows=5)
            for i, col in enumerate(df.columns):
                print(f"{i}: {col}")
        except Exception as e:
            print(f"Error reading with header={header_row}: {e}")
else:
    print("No xlsx files found after conversion.")
